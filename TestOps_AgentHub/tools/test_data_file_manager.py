# tools/test_data_file_manager.py

import requests
import json
import os
import shutil
import glob
from pathlib import Path
from langchain.tools import tool
import pandas as pd

import os
from dotenv import load_dotenv
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))
BASE_URL = os.getenv("HOST_BASE_URL")

# Test case to template mapping
TEST_CASE_MAPPING = {
    "tc_api_supplier_01": "td_oracle_erp_new_supplier_template",
    "bulkapisuppliercreation": "td_oracle_erp_new_supplier_template", 
    "register supplier ui": "td_oracle_erp_new_supplier_template",
    "supplier ui": "td_oracle_erp_new_supplier_template",
    "supplier api": "td_oracle_erp_new_supplier_template",
    "invoice creation ui": "td_oracle_erp_new_invoice_const_amt_template",
    "invoice ui": "td_oracle_erp_new_invoice_const_amt_template"
}

@tool("test_data_file_manager", return_direct=True)
def test_data_file_manager_tool(query: str) -> str:
    """
    Manages test data file operations including updating, adding, and replacing test data files.
    Handles specific templates for Invoice and Supplier test cases with sheet renaming logic.
    """
    try:
        query_lower = query.lower()
        
        # Extract test case name from query - Start directly with main logic
        test_case_name = extract_test_case_name(query_lower)
        
        if not test_case_name:
            return """
❌ **TEST CASE NOT IDENTIFIED**

Please specify a test case name in your request.

**Available test cases:**
- Invoice Creation UI
- Invoice UI  
- BulkAPISupplierCreation
- Register Supplier UI
- Supplier UI
- Supplier API

**Example:** "update test data for Invoice Creation UI"
"""
        
        # Find matching template
        template_name = find_template_for_test_case(test_case_name)
        
        if not template_name:
            return f"""
❌ **TEST DATA NOT AVAILABLE**

Test data is not available for "{test_case_name}" yet.

**Only available for these test cases:**
- TC_API_SUPPLIER_01
- BulkAPISupplierCreation  
- Register Supplier UI
- Supplier UI
- Supplier API
- Invoice Creation UI
- Invoice UI

Please choose one of the above test cases.
"""
        
        # Check if template file exists
        template_path = find_template_file(template_name)
        
        if not template_path:
            return f"""
❌ **TEMPLATE NOT FOUND**

Template "{template_name}" not found in TDM Files location.

You can generate it by asking:
**"Could you please generate test data for template {template_name}"**
"""
        
        # Process the template based on type
        if template_name == "td_oracle_erp_new_invoice_const_amt_template":
            result = process_invoice_template(template_path, test_case_name)
        elif template_name == "td_oracle_erp_new_supplier_template":
            result = process_supplier_template(template_path, test_case_name)
        else:
            return f"❌ Unknown template type: {template_name}"
        
        return result
        
    except Exception as e:
        return f"❌ Error in Test Data File Manager: {str(e)}"

def extract_test_case_name(query: str) -> str:
    """Extract test case name from user query"""
    # Look for "for" keyword to find test case name
    patterns = [
        r"for\s+(.+?)(?:\s*$|\s+and\s|\s+or\s)",
        r"data\s+(.+?)(?:\s*$|\s+and\s|\s+or\s)"
    ]
    
    import re
    for pattern in patterns:
        match = re.search(pattern, query)
        if match:
            return match.group(1).strip()
    
    return ""

def find_template_for_test_case(test_case_name: str) -> str:
    """Find template name for given test case"""
    test_case_lower = test_case_name.lower().strip()
    return TEST_CASE_MAPPING.get(test_case_lower)

def find_template_file(template_name: str) -> str:
    """Find template file in TDM_files folder"""
    # Check current directory structure
    base_dir = os.getcwd()
    tdm_files_path = os.path.join(base_dir, "TDM_files")
    
    if not os.path.exists(tdm_files_path):
        return None
    
    # Look for files that start with template name
    pattern = os.path.join(tdm_files_path, f"{template_name}*")
    matching_files = glob.glob(pattern)
    
    # Filter for Excel files only
    excel_files = [f for f in matching_files if f.endswith(('.xlsx', '.xlsm'))]
    
    return excel_files[0] if excel_files else None

def process_invoice_template(template_path: str, test_case_name: str) -> str:
    """Process invoice template with stage renaming logic"""
    try:
        # Read Excel file to get sheet names
        excel_file = pd.ExcelFile(template_path)
        sheet_names = excel_file.sheet_names
        
        # Find the single stage header and line sheets (whatever stage number they have)
        stage_header = None
        stage_line = None
        
        for sheet_name in sheet_names:
            if 'stage' in sheet_name.lower() and 'header' in sheet_name.lower():
                stage_header = sheet_name
            elif 'stage' in sheet_name.lower() and 'line' in sheet_name.lower():
                stage_line = sheet_name
        
        if not stage_header or not stage_line:
            return f"❌ Could not find stage Header and Line sheets in {os.path.basename(template_path)}. Found sheets: {sheet_names}"
        
        # Create new workbook with renamed sheets
        new_file_path = create_processed_invoice_file_simple(template_path, stage_header, stage_line)
        
        if not new_file_path:
            return f"❌ Failed to process invoice template"
        
        # Upload the processed file with EXACT filename
        upload_result = upload_file_to_endpoint(new_file_path, "invoicedata.xlsx")
        
        # Clean up temporary file
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
        
        if upload_result:
            return f"""
✅ **INVOICE TEST DATA UPDATED SUCCESSFULLY**

🎯 **Test Case**: {test_case_name}
📁 **Template**: td_oracle_erp_new_invoice_const_amt_template
📊 **Processing**: Found {stage_header} & {stage_line} sheets
🔄 **Renamed**: Header & Line sheets created
📤 **Uploaded**: invoicedata.xlsx
🚀 **Status**: Test data file replaced successfully

**Summary**: Invoice test data has been processed and uploaded with stage sheets renamed to Header and Line.
"""
        else:
            return f"❌ Failed to upload processed invoice template"
            
    except Exception as e:
        return f"❌ Error processing invoice template: {str(e)}"

def process_supplier_template(template_path: str, test_case_name: str) -> str:
    """Process supplier template with header renaming logic"""
    try:
        # Read Excel file to get sheet names  
        excel_file = pd.ExcelFile(template_path)
        sheet_names = excel_file.sheet_names
        
        # Find the single stage header sheet (whatever stage number it has)
        stage_header = None
        
        for sheet_name in sheet_names:
            if 'stage' in sheet_name.lower() and 'header' in sheet_name.lower():
                stage_header = sheet_name
                break
        
        if not stage_header:
            return f"❌ Could not find stage Header sheet in {os.path.basename(template_path)}. Found sheets: {sheet_names}"
        
        # Create new workbook with renamed sheet
        new_file_path = create_processed_supplier_file_simple(template_path, stage_header)
        
        if not new_file_path:
            return f"❌ Failed to process supplier template"
        
        # Upload the processed file with EXACT filename
        upload_result = upload_file_to_endpoint(new_file_path, "SupplierImportTemplate.xlsx")
        
        # Clean up temporary file
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
        
        if upload_result:
            return f"""
✅ **SUPPLIER TEST DATA UPDATED SUCCESSFULLY**
🎯 **Test Case**: {test_case_name}
📁 **Template**: td_oracle_erp_new_supplier_template  
📊 **Processing**: Found {stage_header} sheet
🔄 **Renamed**: POZ_SUPPLIERS_INT sheet created
📤 **Uploaded**: SupplierImportTemplate.xlsx
🚀 **Status**: Test data file replaced successfully

**Summary**: Supplier test data has been processed and uploaded with {stage_header} renamed to POZ_SUPPLIERS_INT.
"""
        else:
            return f"❌ Failed to upload processed supplier template"
            
    except Exception as e:
        return f"❌ Error processing supplier template: {str(e)}"

def create_processed_invoice_file_simple(template_path: str, stage_header: str, stage_line: str) -> str:
    """Create processed invoice file with renamed sheets - simplified"""
    temp_dir = os.path.dirname(template_path)
    temp_path = os.path.join(temp_dir, "temp_invoicedata.xlsx")
    
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(template_path)
        
        # Simply rename the two sheets we found
        wb[stage_header].title = 'Headers'
        wb[stage_line].title = 'Lines'
        
        # Save the processed file
        wb.save(temp_path)
        wb.close()
        
        return temp_path
        
    except Exception as e:
        print(f"Error creating processed invoice file: {e}")
        return None

def create_processed_supplier_file_simple(template_path: str, stage_header: str) -> str:
    """Create processed supplier file with renamed sheet - simplified"""
    temp_dir = os.path.dirname(template_path)
    temp_path = os.path.join(temp_dir, "temp_SupplierImportTemplate.xlsx")
    
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(template_path)
        
        # Simply rename the one sheet we found
        wb[stage_header].title = 'POZ_SUPPLIERS_INT'
        
        # Save the processed file
        wb.save(temp_path)
        wb.close()
        
        return temp_path
        
    except Exception as e:
        print(f"Error creating processed supplier file: {e}")
        return None


def find_max_stage(sheet_names: list) -> int:
    """Find the maximum stage number from sheet names"""
    import re
    max_stage = 1
    
    for sheet in sheet_names:
        match = re.search(r'stage(\d+)', sheet.lower())
        if match:
            stage_num = int(match.group(1))
            max_stage = max(max_stage, stage_num)
    
    return max_stage

def create_processed_invoice_file(template_path: str, max_stage: int) -> str:
    """Create processed invoice file with renamed sheets"""
    # Use a simple temporary filename without '_processed'
    temp_dir = os.path.dirname(template_path)
    temp_path = os.path.join(temp_dir, "temp_invoicedata.xlsx")
    
    # Read and modify using openpyxl directly to avoid sheet visibility issues
    from openpyxl import load_workbook
    
    try:
        # Load the workbook
        wb = load_workbook(template_path)
        
        # Find and rename the stage sheets
        sheets_to_remove = []
        
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == f'stage{max_stage}header':
                # Rename to Header
                wb[sheet_name].title = 'Header'
            elif sheet_name.lower() == f'stage{max_stage}line':
                # Rename to Line  
                wb[sheet_name].title = 'Line'
            elif sheet_name.lower().startswith('stage') and ('header' in sheet_name.lower() or 'line' in sheet_name.lower()):
                # Mark other stage sheets for removal
                sheets_to_remove.append(sheet_name)
        
        # Remove unwanted stage sheets
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        
        # Ensure we have at least Header and Line sheets
        if 'Header' not in wb.sheetnames or 'Line' not in wb.sheetnames:
            wb.close()
            return None
        
        # Save the processed file
        wb.save(temp_path)
        wb.close()
        
        return temp_path
        
    except Exception as e:
        print(f"Error creating processed invoice file: {e}")
        return None

def create_processed_supplier_file(template_path: str, max_stage: int) -> str:
    """Create processed supplier file with renamed sheet"""
    # Use a simple temporary filename without '_processed'
    temp_dir = os.path.dirname(template_path)
    temp_path = os.path.join(temp_dir, "temp_SupplierImportTemplate.xlsx")
    
    # Read and modify using openpyxl directly
    from openpyxl import load_workbook
    
    try:
        # Load the workbook
        wb = load_workbook(template_path)
        
        # Find and rename the stage header sheet
        sheets_to_remove = []
        target_sheet_found = False
        
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == f'stage{max_stage}header':
                # Rename to POZ_SUPPLIERS_INT
                wb[sheet_name].title = 'POZ_SUPPLIERS_INT'
                target_sheet_found = True
            elif sheet_name.lower().startswith('stage') and 'header' in sheet_name.lower():
                # Mark other stage headers for removal
                sheets_to_remove.append(sheet_name)
        
        # Remove unwanted stage sheets
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        
        # Ensure we have POZ_SUPPLIERS_INT sheet
        if not target_sheet_found or 'POZ_SUPPLIERS_INT' not in wb.sheetnames:
            wb.close()
            return None
        
        # Save the processed file
        wb.save(temp_path)
        wb.close()
        
        return temp_path
        
    except Exception as e:
        print(f"Error creating processed supplier file: {e}")
        return None

def upload_file_to_endpoint(file_path: str, target_filename: str) -> bool:
    """Upload file using the file upload endpoint with exact target filename"""
    try:
        with open(file_path, 'rb') as file:
            # Use the exact target filename for upload
            files = {'file': (target_filename, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{BASE_URL}/upload-excel", files=files)
            
            if response.status_code == 200:
                print(f"✅ Successfully uploaded: {target_filename}")
                return True
            else:
                print(f"❌ Upload failed: {response.status_code} - {response.text}")
                return False
                
    except Exception as e:
        print(f"Upload error: {e}")
        return False
